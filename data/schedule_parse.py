from datetime import timedelta
import openpyxl
from openpyxl.reader import excel
from data.consts import days_of_the_week
import os

FIRST_SMENA_FILE = 'db\\1_smena_schedule.xlsx'
SECOND_SMENA_FILE = 'db\\2_smena_schedule.xlsx'
LESSONS_EVERY_DAY = 7
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


class ScheduleParser:
    def __init__(self):
        self._excel_to_list()
        self._init_clas_and_paralel_lists()
        self._set_lessons_time()
        self._table_to_dict_schedule()

    def _excel_to_list(self):
        static_root = os.path.join(PROJECT_ROOT, FIRST_SMENA_FILE)
        wb_obj = openpyxl.load_workbook(static_root)
        
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row

        self.hole_table = []

        # Loop will print all columns name
        for j in range(1, max_row + 1):
            self.hole_table.append([])
            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row = j, column = i)
                # print(cell_obj.value)
                self.hole_table[j - 1].append(cell_obj.value)
    
    def _init_clas_and_paralel_lists(self):
        # self.paralel_list - how many classe in each parallel
        # self.clas_list - classes with letters
        self.paralel_list = {str(i): 0 for i in range(8, 11 + 1)}
        self.clas_list = []

        for clas in range(2, len(self.hole_table[0])):
            name = self.hole_table[0][clas]
            if ' ' in name:  # от названия типа "8А (фил / сг)" должно остаться только "8А"
                name = name[: name.index(' ')]

            self.paralel_list[name[:-1]] += 1  # name[:-1] - это номер класса (без буквы)
            self.clas_list.append(name)
    
    def get_paralel_list(self):
        return self.paralel_list

    def _set_lessons_time(self):
        self.week_time_table = [self.hole_table[i][1] for i in range(1, LESSONS_EVERY_DAY + 1)]
        self.saturday_time_table = [
                    self.hole_table[i][1]
                    for i in range(7 * 5 + 2, len(self.hole_table))
                ]  # range is 7 lessons a day, saturday - the 6th day,
                   # thats why multiply 7 by 5 
                   # and plus 2 cuz first row is taken by names of classes
                   # and plus one cuz we start with the next row
        '''
        init_dict = {
            'Понедельник': min_time_table.copy(),
            'Вторник': min_time_table.copy(),
            'Среда': min_time_table.copy(),
            'Четверг': min_time_table.copy(),
            'Пятница': min_time_table.copy(),
            'Суббота': min_time_table.copy()
        } '''
        
    def _get_schedule_for_day(self, clas, day):
        num_of_clas = self.clas_list.index(clas)
        if day == 5:
            timetable = self.saturday_time_table.copy()
            lessons = 6
        else:
            timetable = self.week_time_table.copy()
            lessons = 7
        
        result = {}
        for i in range(lessons):
            if self.hole_table[day * 7 + i + 1][2 + num_of_clas] is not None:
                lesson = self.hole_table[day * 7 + i + 1][2 + num_of_clas]
                lesson_withput_teachers_name = list(lesson.split('\n'))[0]
                result[timetable[i]] = lesson_withput_teachers_name
        return result
    
    def _get_schedule_for_clas(self, clas):
        result = {}
        for day in days_of_the_week.keys():
            result[days_of_the_week[day]] = self._get_schedule_for_day(clas, day)
        return result

    def _table_to_dict_schedule(self):
        # names of classes
        self.dict_schedule = {}
        
        for clas in self.clas_list:
            self.dict_schedule[clas] = self._get_schedule_for_clas(clas)
    
    def get_schedule(self, clas, day):
        return self.dict_schedule[clas][day]
